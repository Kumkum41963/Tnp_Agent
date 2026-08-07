"""
Excel Service — all Excel I/O via OpenPyXL and Pandas.

Responsibilities:
- Read the Master Database Excel/CSV into MasterRecord objects
- Parse raw company uploads into headers + sample rows
- Write the final Populated Company Database back out, preserving the
  company's original template column order and basic formatting
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.models.master_record import MasterRecord


class ExcelServiceError(Exception):
    """Raised on Excel I/O failures."""


# ── Master DB column aliases ─────────────────────────────────────────────────
# Maps common alternative column names in the Master DB file to canonical fields.
MASTER_COLUMN_ALIASES: dict[str, str] = {
    "roll_no": "enrollment_number",
    "roll no": "enrollment_number",
    "enroll_no": "enrollment_number",
    "enrollment no": "enrollment_number",
    "Enrollment No.": "enrollment_number",
    "student_name": "name",
    "full_name": "name",
    "mobile": "phone_number",
    "mobile_number": "phone_number",
    "phone": "phone_number",
    "contact": "phone_number",
    "mail": "email",
    "email_id": "email",
    "department": "branch",
    "gpa": "cgpa",
    "backlogs": "backlog_count",
    "active_backlogs": "backlog_count",
    "dob": "date_of_birth",
    "10th_percent": "tenth_percentage",
    "10th_%": "tenth_percentage",
    "12th_percent": "twelfth_percentage",
    "12th_%": "twelfth_percentage",
    "linkedin": "linkedin_url",
    "github": "github_url",
    "resume": "resume_link",
    "resume_url": "resume_link",
     "enrollment_no.": "enrollment_number",
    "personal_email_id": "email",
    "mobile_number": "phone_number",
    "branch_name": "branch",
    "agreegate_cgpa": "cgpa",

    "xth_percentage": "tenth_percentage",
    "xiith_percentage": "twelfth_percentage",

    "father's_name": "father_name",
    "mother's_name": "mother_name",

    "resume(with_view_access)": "resume_link",

    "active/dead_backlog/__n/a": "backlog_count",
}

KNOWN_MASTER_FIELDS = {
    "enrollment_number", "name", "email", "phone_number", "branch", "cgpa",
    "backlog_count", "gender", "date_of_birth", "address", "father_name",
    "mother_name", "tenth_percentage", "twelfth_percentage", "resume_link",
    "linkedin_url", "github_url",
}


class ExcelService:
    """Handles all Excel / CSV file operations."""

    # ── Master Database ──────────────────────────────────────────────────────

    def load_master_database(self, file_path: str | Path) -> list[MasterRecord]:
        """
        Parse the Master Database Excel/CSV into MasterRecord objects.
        Uses a deterministic column mapping (static aliases above) — no AI involved.
        Raises ExcelServiceError if required columns are missing.
        """
        path = Path(file_path)
        if not path.exists():
            raise ExcelServiceError(f"Master Database file not found: {path}")

        logger.info(f"Loading Master Database from {path}")
        df = self._read_file(path)

        # Normalize column names: lowercase + strip + replace spaces with _
        df.columns = [
            str(c).strip().lower().replace(" ", "_").replace("-", "_")
            for c in df.columns
        ]

        # Apply aliases
        df.rename(
            columns={
                alias: canonical
                for alias, canonical in MASTER_COLUMN_ALIASES.items()
                if alias in df.columns
            },
            inplace=True,
        )

        # Validate required fields i.e no column name actually matches the aliases
        for required in ("enrollment_number", "name"):
            if required not in df.columns:
                raise ExcelServiceError(
                    f"Master Database is missing required column: '{required}'. "
                    f"Available columns: {list(df.columns)}"
                )

        records: list[MasterRecord] = []
        for _, row in df.iterrows():
            known = {
                field: (str(row[field]).strip() if pd.notna(row.get(field)) else None)
                for field in KNOWN_MASTER_FIELDS
                if field in df.columns
            }
            extra = {
                col: (str(row[col]).strip() if pd.notna(row[col]) else None)
                for col in df.columns
                if col not in KNOWN_MASTER_FIELDS
            }
            # Cast numeric fields
            for float_field in ("cgpa", "tenth_percentage", "twelfth_percentage"):
                if float_field in known and known[float_field] is not None:
                    try:
                        known[float_field] = float(known[float_field])  # type: ignore[assignment]
                    except (ValueError, TypeError):
                        known[float_field] = None
            for int_field in ("backlog_count",):
                if int_field in known and known[int_field] is not None:
                    try:
                        known[int_field] = int(float(known[int_field]))  # type: ignore[assignment]
                    except (ValueError, TypeError):
                        known[int_field] = None

            records.append(MasterRecord(**known, extra=extra))

        logger.info(f"Loaded {len(records)} Master Database records.")
        return records

    # ── Company Upload ───────────────────────────────────────────────────────

    def parse_company_upload(
        self, file_path: str | Path
    ) -> tuple[list[str], list[dict[str, str]]]:
        """
        Parse the company Excel file into (headers, sample_rows).

        Returns
        -------
        headers     : List of raw column header strings exactly as the company wrote them.
        sample_rows : Up to 5 rows of raw cell values (as dicts keyed by header),
                      used by the Schema Agent for type disambiguation.
        """
        path = Path(file_path)
        if not path.exists():
            raise ExcelServiceError(f"Company upload not found: {path}")

        df = self._read_file_company(path)
        # Preserve original header capitalisation — do NOT normalise
        headers = [str(c).strip() for c in df.columns]
        df.columns = headers  # type: ignore[assignment]

        sample = df.head(5).fillna("").astype(str)
        sample_rows = sample.to_dict(orient="records")  # type: ignore[arg-type]

        logger.info(
            f"Parsed company upload: {len(headers)} columns, "
            f"{len(df)} rows, {len(sample_rows)} sample rows captured."
        )
        return headers, sample_rows

    # ── Populated Company Database output ────────────────────────────────────

    def write_populated_database(
        self,
        output_path: str | Path,
        company_headers: list[str],
        rows: list[dict[str, str | None]],
    ) -> None:
        """
        Write the populated Company Database to an Excel file, preserving
        the original company column order and applying basic header formatting.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        df = pd.DataFrame(rows, columns=company_headers)

        wb = load_workbook_or_new(output_path)
        ws = wb.active
        ws.title = "Populated Database"  # type: ignore[assignment]
        ws.delete_rows(1, ws.max_row)

        # Write using openpyxl for formatting control
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="2D6A4F", end_color="2D6A4F", fill_type="solid"
                    )

        wb.save(output_path)
        logger.info(f"Written populated database to {output_path}")

    # ── Internal helpers ─────────────────────────────────────────────────────

    @staticmethod
    def _read_file(path: Path) -> pd.DataFrame:
        """Read Excel or CSV into a DataFrame."""
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls", ".xlsm"):
            excel_result = pd.read_excel(path, header=1, dtype=str)
            print(excel_result)
            return excel_result
        elif suffix == ".csv":
            return pd.read_csv(path, dtype=str)
        else:
            raise ExcelServiceError(
                f"Unsupported file format: {suffix}. Expected .xlsx, .xls, or .csv."
            )

    @staticmethod
    def _read_file_company(path: Path) -> pd.DataFrame:
             """Read Excel or CSV of a company db format into a DataFrame."""
             suffix = path.suffix.lower()
             if suffix in (".xlsx", ".xls", ".xlsm"):
                 excel_result = pd.read_excel(path, dtype=str)
                 print(excel_result)
                 return excel_result
             elif suffix == ".csv":
                 return pd.read_csv(path, dtype=str)
             else:
                 raise ExcelServiceError(
                     f"Unsupported file format: {suffix}. Expected .xlsx, .xls, or .csv."
                 )   


def load_workbook_or_new(path: Path):  # type: ignore[return]
    """Load an existing workbook or create a new one if it doesn't exist."""
    from openpyxl import Workbook

    if path.exists():
        return load_workbook(path)
    return Workbook()


# Singleton
excel_service = ExcelService()
